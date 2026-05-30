import argparse
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ZH = {
    "repetition": "重复",
    "short pause": "短暂停顿",
    "long pause": "长停顿",
    "none": "无明显不流畅",
    "hesitation": "犹豫/卡顿",
    "self-correction": "自我修正",
    "minor hesitation before English phrase": "英文短语前有轻微犹豫",
    "slight Mandarin-accented English": "英文带轻微中文口音，但仍可理解",
    "clear Mandarin pronunciation": "普通话发音清晰",
    "clear English pronunciation": "英文发音清晰",
    "clear code-switching": "中英文切换清楚",
    "natural intonation": "语调自然",
    "stable rhythm": "节奏稳定",
    "clipped audio": "音频有爆音/削波，局部失真",
    "partially intelligible": "部分可理解，有些词不够清楚",
    "speech fully intelligible": "语音完全可理解",
    "clear recording": "录音清晰",
    "low background noise": "背景噪声低",
    "consistent volume": "音量稳定",
}


def read_textgrid(path: Path) -> str:
    for encoding in ("utf-16", "utf-8"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeError:
            continue
    return path.read_text()


def parse_tiers(text: str) -> dict[str, list[tuple[float, float, str]]]:
    tiers = {}
    tier_blocks = re.findall(
        r'item \[\d+\]:\s+class = "IntervalTier"\s+name = "(.*?)"\s+xmin = .*?\s+xmax = .*?\s+intervals: size = \d+\s+(.*?)(?=\n    item \[\d+\]:|\Z)',
        text,
        flags=re.S,
    )
    for name, block in tier_blocks:
        rows = []
        for xmin, xmax, label in re.findall(
            r'intervals \[\d+\]:\s+xmin = ([\d.]+)\s+xmax = ([\d.]+)\s+text = "(.*?)"',
            block,
            flags=re.S,
        ):
            rows.append((float(xmin), float(xmax), label.strip()))
        tiers[name] = rows
    return tiers


def label_at(rows: list[tuple[float, float, str]], start: float, end: float) -> str:
    labels = []
    for xmin, xmax, label in rows:
        if not label:
            continue
        overlap = min(xmax, end) - max(xmin, start)
        if overlap > 0.001:
            labels.append(label)
    return "; ".join(dict.fromkeys(labels))


def translate(labels: str) -> str:
    if not labels:
        return ""
    parts = [p.strip() for p in labels.split(";") if p.strip()]
    return "；".join(ZH.get(p, p) for p in parts)


def all_boundaries(tiers: dict[str, list[tuple[float, float, str]]]) -> list[tuple[float, float]]:
    points = set()
    for rows in tiers.values():
        for start, end, label in rows:
            if label:
                points.add(round(start, 6))
                points.add(round(end, 6))
    ordered = sorted(points)
    return [
        (ordered[i], ordered[i + 1])
        for i in range(len(ordered) - 1)
        if ordered[i + 1] > ordered[i]
    ]


def write_xlsx(tiers: dict[str, list[tuple[float, float, str]]], output: Path) -> int:
    headers = [
        "Segment ID",
        "Start Time (s)",
        "End Time (s)",
        "Duration (s)",
        "Transcript",
        "Disfluency Label",
        "Disfluency 中文释义",
        "Pronunciation Label",
        "Pronunciation 中文释义",
        "Audio Quality Label",
        "Audio Quality 中文释义",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Annotation Table"
    ws.append(headers)

    count = 0
    for start, end in all_boundaries(tiers):
        transcript = label_at(tiers.get("Transcript", []), start, end)
        disfluency = label_at(tiers.get("Disfluency", []), start, end)
        pronunciation = label_at(tiers.get("Pronunciation", []), start, end)
        audio_quality = label_at(tiers.get("AudioQuality", []), start, end)
        if not any([transcript, disfluency, pronunciation, audio_quality]):
            continue
        count += 1
        ws.append(
            [
                count,
                round(start, 3),
                round(end, 3),
                round(end - start, 3),
                transcript,
                disfluency,
                translate(disfluency),
                pronunciation,
                translate(pronunciation),
                audio_quality,
                translate(audio_quality),
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [10, 13, 12, 12, 42, 24, 28, 34, 36, 26, 32]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Praat TextGrid annotations to XLSX.")
    parser.add_argument("--input", required=True, help="Input .TextGrid file")
    parser.add_argument("--output", required=True, help="Output .xlsx file")
    args = parser.parse_args()

    textgrid = Path(args.input)
    output = Path(args.output)
    tiers = parse_tiers(read_textgrid(textgrid))
    count = write_xlsx(tiers, output)
    print(f"Created {output} with {count} annotation rows.")


if __name__ == "__main__":
    main()
